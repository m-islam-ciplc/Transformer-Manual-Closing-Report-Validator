"""
Stock Report Validator GUI
Simple GUI for uploading and validating stock reports
"""

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import re
import os

# Color for highlighting matches - light yellow
HIGHLIGHT_FILL = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')

class StockReportValidator:
    def __init__(self, root):
        self.root = root
        self.root.title("Stock Report Validator")
        self.root.geometry("700x600")
        
        self.odoo_file_path = None
        self.manual_file_path = None
        
        self.setup_ui()
    
    def setup_ui(self):
        # Title
        title_label = tk.Label(self.root, text="Stock Report Validator", 
                               font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # File selection frame
        file_frame = tk.Frame(self.root)
        file_frame.pack(pady=20, padx=20, fill=tk.X)
        
        # Odoo file selection
        odoo_frame = tk.Frame(file_frame)
        odoo_frame.pack(fill=tk.X, pady=5)
        tk.Label(odoo_frame, text="Odoo Closing Stock Report:", font=("Arial", 10)).pack(anchor=tk.W)
        odoo_btn_frame = tk.Frame(odoo_frame)
        odoo_btn_frame.pack(fill=tk.X, pady=5)
        tk.Button(odoo_btn_frame, text="Select File 1 (Odoo)", 
                 command=self.select_odoo_file, width=20).pack(side=tk.LEFT)
        self.odoo_label = tk.Label(odoo_btn_frame, text="No file selected", 
                                   fg="gray", anchor=tk.W)
        self.odoo_label.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        
        # Manual file selection
        manual_frame = tk.Frame(file_frame)
        manual_frame.pack(fill=tk.X, pady=5)
        tk.Label(manual_frame, text="Manual Closing Stock Report:", font=("Arial", 10)).pack(anchor=tk.W)
        manual_btn_frame = tk.Frame(manual_frame)
        manual_btn_frame.pack(fill=tk.X, pady=5)
        tk.Button(manual_btn_frame, text="Select File 2 (Manual)", 
                 command=self.select_manual_file, width=20).pack(side=tk.LEFT)
        self.manual_label = tk.Label(manual_btn_frame, text="No file selected", 
                                     fg="gray", anchor=tk.W)
        self.manual_label.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        
        # Validate button
        self.validate_btn = tk.Button(self.root, text="Validate and Match", 
                                     command=self.start_validation,
                                     font=("Arial", 12, "bold"),
                                     bg="#4CAF50", fg="white",
                                     state=tk.DISABLED)
        self.validate_btn.pack(pady=20)
        
        # Progress/Status area
        status_frame = tk.Frame(self.root)
        status_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)
        tk.Label(status_frame, text="Status:", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.status_text = scrolledtext.ScrolledText(status_frame, height=15, wrap=tk.WORD)
        self.status_text.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Update button state
        self.update_button_state()
    
    def select_odoo_file(self):
        filename = filedialog.askopenfilename(
            title="Select Odoo Closing Stock Report",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.odoo_file_path = filename
            self.odoo_label.config(text=os.path.basename(filename), fg="green")
            self.update_button_state()
            self.log_status(f"Selected Odoo file: {os.path.basename(filename)}")
    
    def select_manual_file(self):
        filename = filedialog.askopenfilename(
            title="Select Manual Closing Stock Report",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.manual_file_path = filename
            self.manual_label.config(text=os.path.basename(filename), fg="green")
            self.update_button_state()
            self.log_status(f"Selected Manual file: {os.path.basename(filename)}")
    
    def update_button_state(self):
        if self.odoo_file_path and self.manual_file_path:
            self.validate_btn.config(state=tk.NORMAL)
        else:
            self.validate_btn.config(state=tk.DISABLED)
    
    def log_status(self, message):
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)
        self.root.update()
    
    def start_validation(self):
        """Start validation in a separate thread to keep UI responsive"""
        self.validate_btn.config(state=tk.DISABLED)
        self.status_text.delete(1.0, tk.END)
        
        thread = threading.Thread(target=self.validate_files)
        thread.daemon = True
        thread.start()
    
    def validate_files(self):
        """Main validation logic"""
        try:
            self.log_status("=" * 60)
            self.log_status("Starting validation...")
            self.log_status("=" * 60)
            
            # Read files
            self.log_status("\nReading Odoo file...")
            odoo_data = self.read_odoo_data(self.odoo_file_path)
            self.log_status(f"Found {len(odoo_data)} data rows in Odoo file")
            
            self.log_status("\nReading Manual RM sheet...")
            manual_rm_data = self.read_manual_rm_data(self.manual_file_path)
            self.log_status(f"Found {len(manual_rm_data)} data rows in Manual RM sheet")
            
            self.log_status("\nReading Manual Consumable sheet...")
            manual_consumable_data = self.read_manual_consumable_data(self.manual_file_path)
            self.log_status(f"Found {len(manual_consumable_data)} data rows in Manual Consumable sheet")
            
            # Show all normalized units
            self.log_status("\nCollecting unique units...")
            all_manual_data = manual_rm_data + manual_consumable_data
            self.show_normalized_units(odoo_data, all_manual_data)
            
            # Find matches for RM sheet
            self.log_status("\nFinding matches for RM sheet...")
            rm_matches = self.find_matches(odoo_data, manual_rm_data, prefix='RM')
            self.log_status(f"Found {len(rm_matches)} RM matches")
            
            # Find matches for Consumable sheet
            self.log_status("\nFinding matches for Consumable sheet...")
            consumable_matches = self.find_matches(odoo_data, manual_consumable_data, prefix='CON')
            self.log_status(f"Found {len(consumable_matches)} Consumable matches")
            
            total_matches = len(rm_matches) + len(consumable_matches)
            if total_matches == 0:
                self.log_status("\nNo matches found!")
                messagebox.showinfo("No Matches", "No matching records found between the two files.")
                self.validate_btn.config(state=tk.NORMAL)
                return
            
            # Process files
            self.log_status("\nProcessing files...")
            self.process_files(self.odoo_file_path, self.manual_file_path, rm_matches, consumable_matches)
            
            # Generate analysis report in the same folder as loaded files
            self.log_status("\nGenerating analysis report...")
            import os
            report_dir = os.path.dirname(self.manual_file_path) or os.path.dirname(self.odoo_file_path) or '.'
            all_matches = rm_matches + consumable_matches
            all_manual_data = manual_rm_data + manual_consumable_data
            self.generate_analysis_report(odoo_data, all_manual_data, all_matches, rm_matches, consumable_matches, report_dir)
            
            self.log_status("\n" + "=" * 60)
            self.log_status("VALIDATION COMPLETED SUCCESSFULLY!")
            self.log_status("=" * 60)
            self.log_status(f"\nTotal matches found: {total_matches}")
            self.log_status(f"  - RM matches: {len(rm_matches)}")
            self.log_status(f"  - Consumable matches: {len(consumable_matches)}")
            self.log_status("\nMatch Summary (first 10):")
            for match in all_matches[:10]:
                name_preview = match['name'][:40] + "..." if len(match['name']) > 40 else match['name']
                self.log_status(f"  {match['match_id']}: Code='{match['product_code']}', Name='{name_preview}'")
            if total_matches > 10:
                self.log_status(f"  ... and {total_matches - 10} more matches")
            
            messagebox.showinfo("Success", 
                              f"Validation completed!\n\nFound {total_matches} matches:\n"
                              f"- RM: {len(rm_matches)} matches\n"
                              f"- Consumable: {len(consumable_matches)} matches\n\n"
                              "Both files have been updated with:\n"
                              "- Match IDs in the first column\n"
                              "- Highlighted matching rows in light yellow\n\n"
                              "Analysis report saved in the same folder as loaded files.")
            
        except Exception as e:
            error_msg = f"Error during validation: {str(e)}"
            self.log_status(f"\nERROR: {error_msg}")
            messagebox.showerror("Error", error_msg)
        finally:
            self.validate_btn.config(state=tk.NORMAL)
    
    def is_data_row_odoo(self, row_vals):
        """Check if a row is a data row in Odoo file"""
        if not row_vals or len(row_vals) == 0:
            return False
        sl_val = str(row_vals[0]).strip() if row_vals[0] else ''
        return bool(re.match(r'^\d+$', sl_val))
    
    def is_data_row_manual(self, row_vals):
        """Check if a row is a data row in Manual file"""
        if not row_vals or len(row_vals) < 4:
            return False
        sl_val = str(row_vals[0]).strip() if row_vals[0] else ''
        if not re.match(r'^\d+$', sl_val):
            return False
        product_code = str(row_vals[1]).strip() if len(row_vals) > 1 and row_vals[1] else ''
        items_name = str(row_vals[3]).strip() if len(row_vals) > 3 and row_vals[3] else ''
        return bool(product_code or items_name)
    
    def normalize_text(self, text):
        """Normalize text for comparison"""
        if text is None:
            return ''
        return str(text).strip()
    
    def normalize_product_code(self, text):
        """Normalize product code - remove all spaces and make case-insensitive for comparison"""
        if text is None:
            return ''
        # Remove all spaces (leading, trailing, and internal) and convert to uppercase
        return str(text).replace(' ', '').upper()
    
    def normalize_unit(self, text):
        """Normalize unit - case-insensitive comparison with unit aliases"""
        if text is None:
            return ''
        
        # Normalize to uppercase and strip whitespace
        unit = str(text).strip().upper()
        
        # Map "PCS", "Pieces", and "Piece" variations to "PCS"
        if unit in ('PCS', 'PIECES', 'PIECE'):
            return 'PCS'
        
        # Map "Foot(ft)" and "Feet" variations to "FEET"
        # Handle "FOOT(FT)", "FOOT (FT)", "FEET", etc.
        if unit in ('FOOT(FT)', 'FOOT (FT)', 'FOOT', 'FEET'):
            return 'FEET'
        
        # Map "Liter(s)" and "Liter" variations to "LITER"
        # Handle "LITER(S)", "LITER (S)", "LITER", "LITERS", etc.
        if unit in ('LITER(S)', 'LITER (S)', 'LITER', 'LITERS', 'LITRE(S)', 'LITRE (S)', 'LITRE', 'LITRES'):
            return 'LITER'
        
        # Map "gal(s)" and "Gallon" variations to "GALLON"
        # Handle "GAL(S)", "GAL (S)", "GAL", "GALLON", "GALLONS", etc.
        if unit in ('GAL(S)', 'GAL (S)', 'GAL', 'GALLON', 'GALLONS'):
            return 'GALLON'
        
        # Map "Square Foot" and "SFT" variations to "SFT"
        # Handle "SQUARE FOOT", "SQUARE FEET", "SFT", "SQ FT", "SQFT", etc.
        if unit in ('SQUARE FOOT', 'SQUARE FEET', 'SQUARE FOOT(FT)', 'SQUARE FOOT (FT)', 'SFT', 'SQ FT', 'SQFT', 'SQ.FT', 'SQ. FT'):
            return 'SFT'
        
        # Map "lbs" and "Pound" variations to "POUND"
        # Handle "LBS", "LB", "POUND", "POUNDS", etc.
        if unit in ('LBS', 'LB', 'LBS.', 'LB.', 'POUND', 'POUNDS'):
            return 'POUND'
        
        # Map "Meter" variations to "METER"
        # Handle "METER", "METERS", "METRE", "METRES", "MTRS", "MTR", "MITER", etc.
        if unit in ('METER', 'METERS', 'METRE', 'METRES', 'MTRS', 'MTR', 'MTR.', 'MITER'):
            return 'METER'
        
        # Map "Ream" and "Rim" variations to "REAM"
        # Handle "REAM", "REAMS", "RIM", "RIMS", etc.
        if unit in ('REAM', 'REAMS', 'RIM', 'RIMS'):
            return 'REAM'
        
        return unit
    
    def show_normalized_units(self, odoo_data, manual_data):
        """Collect and display all unique units with their normalized forms"""
        units_map = {}  # normalized_unit -> set of original units
        
        # Collect units from Odoo data
        for row in odoo_data:
            original_unit = str(row.get('unit', '') or '').strip()
            if original_unit:
                normalized = self.normalize_unit(original_unit)
                if normalized not in units_map:
                    units_map[normalized] = set()
                units_map[normalized].add(original_unit)
        
        # Collect units from Manual data
        for row in manual_data:
            original_unit = str(row.get('unit', '') or '').strip()
            if original_unit:
                normalized = self.normalize_unit(original_unit)
                if normalized not in units_map:
                    units_map[normalized] = set()
                units_map[normalized].add(original_unit)
        
        # Display results
        if units_map:
            self.log_status(f"\nFound {len(units_map)} unique normalized units:")
            self.log_status("-" * 60)
            # Sort by normalized unit for consistent display
            for normalized_unit in sorted(units_map.keys()):
                original_units = sorted(units_map[normalized_unit])
                if len(original_units) == 1 and original_units[0].upper() == normalized_unit:
                    # No transformation needed
                    self.log_status(f"  '{normalized_unit}'")
                else:
                    # Show mapping
                    originals_str = "', '".join(original_units)
                    self.log_status(f"  '{normalized_unit}' ← ['{originals_str}']")
            self.log_status("-" * 60)
        else:
            self.log_status("No units found in data.")
    
    def read_odoo_data(self, file_path):
        """Read all data rows from Odoo file"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Check if Match ID column exists
        has_match_id = str(ws.cell(5, 1).value or '').strip() == 'Match ID'
        
        data_rows = []
        
        # Find all header rows
        header_rows = [5]
        sl_col = 2 if has_match_id else 1
        for i in range(6, ws.max_row + 1):
            first_cell = str(ws.cell(i, sl_col).value or '').strip()
            if first_cell == 'SL\nNo' or first_cell == 'SL No' or first_cell == 'SL':
                header_rows.append(i)
        
        # Adjust column indices if Match ID column exists
        if has_match_id:
            code_col = 3
            name_col = 4
            unit_col = 6
            opening_qty_col = 7
            opening_value_col = 8
            receive_qty_col = 9
            receive_value_col = 10
            issue_qty_col = 11
            issue_value_col = 12
            closing_qty_col = 13
            closing_value_col = 14
        else:
            code_col = 2
            name_col = 3
            unit_col = 5
            opening_qty_col = 6
            opening_value_col = 7
            receive_qty_col = 8
            receive_value_col = 9
            issue_qty_col = 10
            issue_value_col = 11
            closing_qty_col = 12
            closing_value_col = 13
        
        # Read data from each page
        for page_idx in range(len(header_rows)):
            header_start = header_rows[page_idx]
            header_next = header_rows[page_idx + 1] if page_idx + 1 < len(header_rows) else ws.max_row + 1
            
            for row_num in range(header_start + 1, header_next):
                sl_val = ws.cell(row_num, sl_col).value
                # Check if it's a data row (has numeric SL)
                if sl_val is None:
                    continue
                sl_str = str(sl_val).strip()
                if not (sl_str.isdigit() or (isinstance(sl_val, (int, float)))):
                    continue
                
                # Read values using correct column indices
                product_code = self.normalize_text(ws.cell(row_num, code_col).value)
                product_name = self.normalize_text(ws.cell(row_num, name_col).value)
                unit = self.normalize_text(ws.cell(row_num, unit_col).value)
                opening_qty = ws.cell(row_num, opening_qty_col).value
                opening_value = ws.cell(row_num, opening_value_col).value
                receive_qty = ws.cell(row_num, receive_qty_col).value
                receive_value = ws.cell(row_num, receive_value_col).value
                issue_qty = ws.cell(row_num, issue_qty_col).value
                issue_value = ws.cell(row_num, issue_value_col).value
                closing_qty = ws.cell(row_num, closing_qty_col).value
                closing_value = ws.cell(row_num, closing_value_col).value
                
                if product_code or product_name:
                    data_rows.append({
                        'row_num': row_num,
                        'product_code': product_code,
                        'product_name': product_name,
                        'unit': unit,
                        'opening_qty': opening_qty,
                        'opening_value': opening_value,
                        'receive_qty': receive_qty,
                        'receive_value': receive_value,
                        'issue_qty': issue_qty,
                        'issue_value': issue_value,
                        'closing_qty': closing_qty,
                        'closing_value': closing_value
                    })
        
        wb.close()
        return data_rows
    
    def read_manual_rm_data(self, file_path):
        """Read all data rows from Manual RM sheet
        Uses data_only=True to get calculated values from formulas for comparison"""
        # Check if file has Match ID column (already processed)
        wb_check = openpyxl.load_workbook(file_path, data_only=True)
        ws_check = wb_check['RM']
        has_match_id = str(ws_check.cell(5, 1).value or '').strip() == 'Match ID'
        wb_check.close()
        
        # Load with data_only=True to get calculated values (not formulas)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['RM']
        
        data_rows = []
        
        # Adjust column indices if Match ID column exists
        if has_match_id:
            sl_col = 2
            code_col = 3
            name_col = 5
            unit_col = 6
            opening_qty_col = 7
            opening_value_col = 8
            receive_qty_col = 9
            receive_value_col = 10
            issue_qty_col = 11
            issue_value_col = 12
            closing_qty_col = 13
            closing_value_col = 14
        else:
            sl_col = 1
            code_col = 2
            name_col = 4
            unit_col = 5
            opening_qty_col = 6
            opening_value_col = 7
            receive_qty_col = 8
            receive_value_col = 9
            issue_qty_col = 10
            issue_value_col = 11
            closing_qty_col = 12
            closing_value_col = 13
        
        for row_num in range(6, ws.max_row + 1):
            sl_val = ws.cell(row_num, sl_col).value
            # Check if it's a data row (has numeric SL)
            if sl_val is None:
                continue
            sl_str = str(sl_val).strip()
            if not (sl_str.isdigit() or (isinstance(sl_val, (int, float)))):
                continue
            
            # Read values using correct column indices
            product_code = self.normalize_text(ws.cell(row_num, code_col).value)
            items_name = self.normalize_text(ws.cell(row_num, name_col).value)
            unit = self.normalize_text(ws.cell(row_num, unit_col).value)
            opening_qty = ws.cell(row_num, opening_qty_col).value
            opening_value = ws.cell(row_num, opening_value_col).value
            receive_qty = ws.cell(row_num, receive_qty_col).value
            receive_value = ws.cell(row_num, receive_value_col).value
            issue_qty = ws.cell(row_num, issue_qty_col).value
            issue_value = ws.cell(row_num, issue_value_col).value
            closing_qty = ws.cell(row_num, closing_qty_col).value
            closing_value = ws.cell(row_num, closing_value_col).value
            
            if product_code or items_name:
                data_rows.append({
                    'row_num': row_num,
                    'product_code': product_code,
                    'items_name': items_name,
                    'unit': unit,
                    'opening_qty': opening_qty,
                    'opening_value': opening_value,
                    'receive_qty': receive_qty,
                    'receive_value': receive_value,
                    'issue_qty': issue_qty,
                    'issue_value': issue_value,
                    'closing_qty': closing_qty,
                    'closing_value': closing_value
                })
        
        wb.close()
        return data_rows
    
    def read_manual_consumable_data(self, file_path):
        """Read all data rows from Manual Consumable sheet
        Uses data_only=True to get calculated values from formulas for comparison"""
        # Check if file has Match ID column (already processed)
        wb_check = openpyxl.load_workbook(file_path, data_only=True)
        if 'Consumable' not in wb_check.sheetnames:
            wb_check.close()
            return []  # Consumable sheet doesn't exist
        ws_check = wb_check['Consumable']
        has_match_id = str(ws_check.cell(5, 1).value or '').strip() == 'Match ID'
        wb_check.close()
        
        # Load with data_only=True to get calculated values (not formulas)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['Consumable']
        
        data_rows = []
        
        # Adjust column indices if Match ID column exists
        if has_match_id:
            sl_col = 2
            code_col = 3
            name_col = 5
            unit_col = 6
            opening_qty_col = 7
            opening_value_col = 8
            receive_qty_col = 9
            receive_value_col = 10
            issue_qty_col = 11
            issue_value_col = 12
            closing_qty_col = 13
            closing_value_col = 14
        else:
            sl_col = 1
            code_col = 2
            name_col = 4
            unit_col = 5
            opening_qty_col = 6
            opening_value_col = 7
            receive_qty_col = 8
            receive_value_col = 9
            issue_qty_col = 10
            issue_value_col = 11
            closing_qty_col = 12
            closing_value_col = 13
        
        for row_num in range(6, ws.max_row + 1):
            sl_val = ws.cell(row_num, sl_col).value
            # Check if it's a data row (has numeric SL)
            if sl_val is None:
                continue
            sl_str = str(sl_val).strip()
            if not (sl_str.isdigit() or (isinstance(sl_val, (int, float)))):
                continue
            
            # Read values using correct column indices
            product_code = self.normalize_text(ws.cell(row_num, code_col).value)
            items_name = self.normalize_text(ws.cell(row_num, name_col).value)
            unit = self.normalize_text(ws.cell(row_num, unit_col).value)
            opening_qty = ws.cell(row_num, opening_qty_col).value
            opening_value = ws.cell(row_num, opening_value_col).value
            receive_qty = ws.cell(row_num, receive_qty_col).value
            receive_value = ws.cell(row_num, receive_value_col).value
            issue_qty = ws.cell(row_num, issue_qty_col).value
            issue_value = ws.cell(row_num, issue_value_col).value
            closing_qty = ws.cell(row_num, closing_qty_col).value
            closing_value = ws.cell(row_num, closing_value_col).value
            
            if product_code or items_name:
                data_rows.append({
                    'row_num': row_num,
                    'product_code': product_code,
                    'items_name': items_name,
                    'unit': unit,
                    'opening_qty': opening_qty,
                    'opening_value': opening_value,
                    'receive_qty': receive_qty,
                    'receive_value': receive_value,
                    'issue_qty': issue_qty,
                    'issue_value': issue_value,
                    'closing_qty': closing_qty,
                    'closing_value': closing_value
                })
        
        wb.close()
        return data_rows
    
    def normalize_numeric(self, text):
        """Normalize numeric values for exact comparison - handles Accounting format (commas, parentheses) and Number format"""
        if text is None:
            return 0.0
        
        # If it's already a number (int or float), convert directly
        if isinstance(text, (int, float)):
            return round(float(text), 2)
        
        # Convert to string and clean up
        text_str = str(text).strip()
        if text_str == '' or text_str.lower() == 'none':
            return 0.0
        
        # Handle dash/hyphen as zero (common in Accounting format)
        if text_str == '-' or text_str == '—' or text_str == '–':
            return 0.0
        
        # Handle Accounting format: parentheses indicate negative numbers
        is_negative = False
        if text_str.startswith('(') and text_str.endswith(')'):
            is_negative = True
            text_str = text_str[1:-1].strip()  # Remove parentheses
        
        # Remove spaces, commas (thousands separators), currency symbols, and other formatting
        text_str = text_str.replace(' ', '').replace(',', '').replace('$', '').replace('₹', '').replace('€', '').replace('£', '')
        
        if text_str == '' or text_str == '-':
            return 0.0
        
        try:
            # Convert to float
            num_value = float(text_str)
            # Apply negative sign if parentheses were present
            if is_negative:
                num_value = -num_value
            # Round to 2 decimal places for exact comparison
            return round(num_value, 2)
        except (ValueError, TypeError):
            # If conversion fails, return 0.0
            return 0.0
    
    def compare_numeric(self, val1, val2):
        """Compare two numeric values exactly (no tolerance)"""
        num1 = self.normalize_numeric(val1)
        num2 = self.normalize_numeric(val2)
        return num1 == num2
    
    def find_matches(self, odoo_data, manual_data, prefix='RM'):
        """Find matching rows with specified prefix (RM or CON)
        Matches on: Product Code, Product Name, Unit, and all Qty/Value fields"""
        matches = []
        match_id_counter = 1
        
        # Debug: Log first few rows from each file
        if odoo_data:
            self.log_status(f"\nSample Odoo row: Code='{odoo_data[0].get('product_code', '')}', Name='{odoo_data[0].get('product_name', '')[:30]}...', Unit='{odoo_data[0].get('unit', '')}', Opening Qty='{odoo_data[0].get('opening_qty', '')}'")
        if manual_data:
            self.log_status(f"Sample Manual row: Code='{manual_data[0].get('product_code', '')}', Name='{manual_data[0].get('items_name', '')[:30]}...', Unit='{manual_data[0].get('unit', '')}', Opening Qty='{manual_data[0].get('opening_qty', '')}'")
        
        for odoo_row in odoo_data:
            for manual_row in manual_data:
                # Check all matching criteria
                # Normalize product codes (ignore spaces)
                odoo_code_normalized = self.normalize_product_code(odoo_row['product_code'])
                manual_code_normalized = self.normalize_product_code(manual_row['product_code'])
                product_code_match = (odoo_code_normalized and manual_code_normalized and 
                                     odoo_code_normalized == manual_code_normalized)
                
                if not product_code_match:
                    continue  # Skip if product code doesn't match
                
                product_name_match = (odoo_row['product_name'] and manual_row['items_name'] and 
                                     odoo_row['product_name'] == manual_row['items_name'])
                
                if not product_name_match:
                    continue  # Skip if product name doesn't match
                
                # Normalize units (case-insensitive)
                odoo_unit_normalized = self.normalize_unit(odoo_row['unit'])
                manual_unit_normalized = self.normalize_unit(manual_row['unit'])
                unit_match = (odoo_unit_normalized and manual_unit_normalized and 
                             odoo_unit_normalized == manual_unit_normalized)
                
                if not unit_match:
                    continue  # Skip if unit doesn't match
                
                # Compare numeric values with tolerance for floating point precision
                opening_qty_match = self.compare_numeric(odoo_row['opening_qty'], manual_row['opening_qty'])
                opening_value_match = self.compare_numeric(odoo_row['opening_value'], manual_row['opening_value'])
                receive_qty_match = self.compare_numeric(odoo_row['receive_qty'], manual_row['receive_qty'])
                receive_value_match = self.compare_numeric(odoo_row['receive_value'], manual_row['receive_value'])
                issue_qty_match = self.compare_numeric(odoo_row['issue_qty'], manual_row['issue_qty'])
                issue_value_match = self.compare_numeric(odoo_row['issue_value'], manual_row['issue_value'])
                closing_qty_match = self.compare_numeric(odoo_row['closing_qty'], manual_row['closing_qty'])
                closing_value_match = self.compare_numeric(odoo_row['closing_value'], manual_row['closing_value'])
                
                # Debug: Log why a potential match failed (only for first few attempts)
                if product_code_match and product_name_match and match_id_counter <= 3:
                    failed_fields = []
                    if not unit_match:
                        failed_fields.append(f"Unit (Odoo: '{odoo_row['unit']}' vs Manual: '{manual_row['unit']}')")
                    if not opening_qty_match:
                        failed_fields.append(f"Opening Qty (Odoo: '{odoo_row['opening_qty']}' vs Manual: '{manual_row['opening_qty']}')")
                    if not opening_value_match:
                        failed_fields.append(f"Opening Value (Odoo: '{odoo_row['opening_value']}' vs Manual: '{manual_row['opening_value']}')")
                    if failed_fields:
                        self.log_status(f"Potential match failed on: {', '.join(failed_fields[:3])}")
                
                # All criteria must match
                if (product_code_match and product_name_match and unit_match and
                    opening_qty_match and opening_value_match and
                    receive_qty_match and receive_value_match and
                    issue_qty_match and issue_value_match and
                    closing_qty_match and closing_value_match):
                    
                    # Use specified prefix (RM or CON)
                    match_id = f'{prefix}{str(match_id_counter).zfill(4)}'
                    matches.append({
                        'match_id': match_id,
                        'odoo_row_num': odoo_row['row_num'],
                        'manual_row_num': manual_row['row_num'],
                        'product_code': odoo_row['product_code'],
                        'name': odoo_row['product_name']
                    })
                    match_id_counter += 1
                    break
        
        return matches
    
    def adjust_formulas_after_insert(self, ws, inserted_col=1):
        """Adjust formulas after inserting a column - shift column references right by 1"""
        self.log_status("Adjusting formulas...")
        formula_count = 0
        
        # Pattern to match Excel column references (like FN6, $FN$6, FN$6, $FN6)
        col_pattern = re.compile(r'(\$?)([A-Z]+)(\$?)(\d+)')
        
        def shift_column(match):
            """Shift column reference by inserted_col positions"""
            dollar_before = match.group(1)
            col_letters = match.group(2)
            dollar_after = match.group(3)
            row_num = match.group(4)
            
            # Convert column letters to number (A=1, B=2, ..., Z=26, AA=27, etc.)
            col_num = 0
            for char in col_letters:
                col_num = col_num * 26 + (ord(char) - ord('A') + 1)
            
            # Shift column
            new_col_num = col_num + inserted_col
            
            # Convert back to letters
            new_col_letters = ''
            temp = new_col_num
            while temp > 0:
                temp -= 1
                new_col_letters = chr(ord('A') + (temp % 26)) + new_col_letters
                temp //= 26
            
            return f"{dollar_before}{new_col_letters}{dollar_after}{row_num}"
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:  # Formula cell
                    try:
                        formula = str(cell.value)
                        # Adjust column references in formula
                        new_formula = col_pattern.sub(shift_column, formula)
                        cell.value = new_formula
                        formula_count += 1
                    except Exception as e:
                        # If adjustment fails, log warning but keep original
                        self.log_status(f"Warning: Could not adjust formula in {cell.coordinate}")
        
        self.log_status(f"Adjusted {formula_count} formulas")
    
    def clean_odoo_file(self, odoo_wb):
        """Remove blank rows and images from Odoo workbook"""
        odoo_ws = odoo_wb.active
        
        # Remove images from all sheets
        total_images = 0
        for sheet in odoo_wb.worksheets:
            if hasattr(sheet, '_images'):
                image_count = len(sheet._images)
                sheet._images = []
                total_images += image_count
        if total_images > 0:
            self.log_status(f"  Removed {total_images} image(s)")
        
        # Remove blank rows (rows with no data)
        # Start from the bottom to avoid index shifting issues
        rows_to_delete = []
        max_row = odoo_ws.max_row
        
        # Find all header rows first (to preserve them)
        header_rows = set()
        for row_num in range(1, max_row + 1):
            first_cell = str(odoo_ws.cell(row_num, 1).value or '').strip()
            if first_cell == 'SL\nNo' or first_cell == 'SL No' or first_cell == 'Match ID':
                header_rows.add(row_num)
        
        # Check each row (skip header rows and first 5 rows which might contain title/header info)
        for row_num in range(max_row, 5, -1):
            if row_num in header_rows:
                continue  # Skip header rows
            
            # Check if row is empty (all cells are None or empty)
            is_empty = True
            for col in range(1, odoo_ws.max_column + 1):
                cell_value = odoo_ws.cell(row_num, col).value
                if cell_value is not None:
                    # Check if it's not just whitespace
                    if isinstance(cell_value, str) and cell_value.strip():
                        is_empty = False
                        break
                    elif not isinstance(cell_value, str):
                        is_empty = False
                        break
            
            if is_empty:
                rows_to_delete.append(row_num)
        
        # Delete blank rows
        if rows_to_delete:
            for row_num in rows_to_delete:
                odoo_ws.delete_rows(row_num)
            self.log_status(f"  Removed {len(rows_to_delete)} blank row(s)")
    
    def copy_cell_format(self, source_cell, target_cell):
        """Copy font formatting from source cell to target cell"""
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )
    
    def process_sheet(self, odoo_wb, manual_wb, sheet_name, matches, is_first_sheet=False):
        """Process a single sheet (RM or Consumable)"""
        odoo_ws = odoo_wb.active
        
        if sheet_name not in manual_wb.sheetnames:
            return  # Sheet doesn't exist, skip
        
        manual_ws = manual_wb[sheet_name]
        
        # Insert Match ID column in Odoo file only for first sheet
        if is_first_sheet:
            self.log_status(f"Inserting Match ID column in Odoo file...")
            # Preserve column widths for Odoo file
            odoo_col_widths = {}
            default_width = odoo_ws.column_dimensions.group_width if hasattr(odoo_ws.column_dimensions, 'group_width') else None
            default_width = default_width if default_width else 8.43  # Excel default column width
            
            for col_idx in range(1, odoo_ws.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if col_letter in odoo_ws.column_dimensions and odoo_ws.column_dimensions[col_letter].width:
                    odoo_col_widths[col_idx] = odoo_ws.column_dimensions[col_letter].width
                else:
                    odoo_col_widths[col_idx] = default_width
            
            odoo_ws.insert_cols(1)
            
            # Restore ALL column widths (shifted by 1)
            for col_idx, width in odoo_col_widths.items():
                new_col_idx = col_idx + 1
                new_col_letter = openpyxl.utils.get_column_letter(new_col_idx)
                odoo_ws.column_dimensions[new_col_letter].width = width
            
            # Set width for new Match ID column
            match_id_width = odoo_col_widths.get(1, 12.0)
            odoo_ws.column_dimensions['A'].width = max(match_id_width, 12.0)
            
            # Copy font format from existing header cell (row 5, col 2) to new Match ID header
            if odoo_ws.cell(5, 2).value is not None:
                self.copy_cell_format(odoo_ws.cell(5, 2), odoo_ws.cell(5, 1))
            odoo_ws.cell(5, 1).value = 'Match ID'
        
        # Insert Match ID column in Manual sheet
        self.log_status(f"Inserting Match ID column in {sheet_name} sheet...")
        
        # Preserve column widths for Manual file
        # Read ALL column widths (including defaults for columns without explicit widths)
        manual_col_widths = {}
        default_width = manual_ws.column_dimensions.group_width if hasattr(manual_ws.column_dimensions, 'group_width') else None
        default_width = default_width if default_width else 8.43  # Excel default column width
        
        for col_idx in range(1, manual_ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if col_letter in manual_ws.column_dimensions and manual_ws.column_dimensions[col_letter].width:
                # Column has explicit width
                manual_col_widths[col_idx] = manual_ws.column_dimensions[col_letter].width
            else:
                # Column uses default width
                manual_col_widths[col_idx] = default_width
        
        manual_ws.insert_cols(1)
        
        # Restore ALL column widths (shifted by 1)
        for col_idx, width in manual_col_widths.items():
            new_col_idx = col_idx + 1
            new_col_letter = openpyxl.utils.get_column_letter(new_col_idx)
            manual_ws.column_dimensions[new_col_letter].width = width
        
        # Set width for new Match ID column (use width of original column A or default)
        match_id_width = manual_col_widths.get(1, 12.0)
        manual_ws.column_dimensions['A'].width = max(match_id_width, 12.0)  # At least 12 for match IDs
        
        # Copy font format from existing header cell (row 5, col 2) to new Match ID header
        if manual_ws.cell(5, 2).value is not None:
            self.copy_cell_format(manual_ws.cell(5, 2), manual_ws.cell(5, 1))
        manual_ws.cell(5, 1).value = 'Match ID'
        
        # Adjust formulas in Manual file (which has formulas)
        self.adjust_formulas_after_insert(manual_ws, inserted_col=1)
        
        # Apply match IDs and highlight
        self.log_status(f"Applying match IDs and highlighting rows for {sheet_name} sheet...")
        for match in matches:
            # Odoo file
            odoo_row = match['odoo_row_num']
            odoo_ws.cell(odoo_row, 1).value = match['match_id']
            for col in range(1, odoo_ws.max_column + 1):
                odoo_ws.cell(odoo_row, col).fill = HIGHLIGHT_FILL
            
            # Manual file
            manual_row = match['manual_row_num']
            manual_ws.cell(manual_row, 1).value = match['match_id']
            # Highlight columns 1-14 (Match ID + original columns A-M, excluding N)
            for col in range(1, 15):
                manual_ws.cell(manual_row, col).fill = HIGHLIGHT_FILL
    
    def process_files(self, odoo_file, manual_file, rm_matches, consumable_matches):
        """Process and update files for both RM and Consumable sheets"""
        # Load files with formulas preserved (data_only=False is default)
        self.log_status("Loading files (preserving formulas)...")
        odoo_wb = openpyxl.load_workbook(odoo_file, data_only=False)
        manual_wb = openpyxl.load_workbook(manual_file, data_only=False)
        
        # Process RM sheet (first sheet - inserts Odoo Match ID column)
        if rm_matches:
            self.process_sheet(odoo_wb, manual_wb, 'RM', rm_matches, is_first_sheet=True)
        
        # Process Consumable sheet (second sheet - Odoo Match ID already exists)
        if consumable_matches:
            self.process_sheet(odoo_wb, manual_wb, 'Consumable', consumable_matches, is_first_sheet=False)
        
        # Clean Odoo file (remove blank rows and images)
        self.log_status("Cleaning Odoo file (removing blank rows and images)...")
        self.clean_odoo_file(odoo_wb)
        
        # Save files
        self.log_status("Saving files...")
        odoo_wb.save(odoo_file)
        manual_wb.save(manual_file)
        
        odoo_wb.close()
        manual_wb.close()
    
    def generate_analysis_report(self, odoo_data, manual_data, matches, rm_matches, consumable_matches, output_dir='.'):
        """Generate analysis report of unmatched records
        Args:
            odoo_data: List of Odoo data rows
            manual_data: List of Manual data rows
            matches: List of all matched records (RM + Consumable)
            rm_matches: List of RM matched records
            consumable_matches: List of Consumable matched records
            output_dir: Directory where to save the report (default: current directory)
        """
        import os
        from datetime import datetime
        
        # Use the provided output directory (same folder as loaded files)
        report_path = os.path.join(output_dir, 'match_analysis_report.txt')
        
        # Get matched product codes and names
        matched_codes = {m['product_code'] for m in matches}
        matched_names = {m['name'] for m in matches}
        
        # Find unmatched records
        unmatched_odoo = []
        unmatched_manual = []
        name_matches_but_different = []
        
        for odoo_row in odoo_data:
            odoo_code_norm = self.normalize_product_code(odoo_row['product_code'])
            odoo_name = odoo_row['product_name']
            
            # Check if this record was matched
            is_matched = False
            for match in matches:
                if (self.normalize_product_code(match['product_code']) == odoo_code_norm and
                    match['name'] == odoo_name):
                    is_matched = True
                    break
            
            if not is_matched:
                # Check if there's a name match but other differences
                name_match_found = False
                for manual_row in manual_data:
                    if odoo_name == manual_row['items_name']:
                        name_match_found = True
                        # Check differences
                        differences = []
                        if self.normalize_product_code(odoo_row['product_code']) != self.normalize_product_code(manual_row['product_code']):
                            differences.append(f"Product Code: '{odoo_row['product_code']}' vs '{manual_row['product_code']}'")
                        if self.normalize_unit(odoo_row['unit']) != self.normalize_unit(manual_row['unit']):
                            differences.append(f"Unit: '{odoo_row['unit']}' vs '{manual_row['unit']}'")
                        if self.compare_numeric(odoo_row['opening_qty'], manual_row['opening_qty']) == False:
                            differences.append(f"Opening Qty: {odoo_row['opening_qty']} vs {manual_row['opening_qty']}")
                        if self.compare_numeric(odoo_row['opening_value'], manual_row['opening_value']) == False:
                            differences.append(f"Opening Value: {odoo_row['opening_value']} vs {manual_row['opening_value']}")
                        if self.compare_numeric(odoo_row['receive_qty'], manual_row['receive_qty']) == False:
                            differences.append(f"Receive Qty: {odoo_row['receive_qty']} vs {manual_row['receive_qty']}")
                        if self.compare_numeric(odoo_row['receive_value'], manual_row['receive_value']) == False:
                            differences.append(f"Receive Value: {odoo_row['receive_value']} vs {manual_row['receive_value']}")
                        if self.compare_numeric(odoo_row['issue_qty'], manual_row['issue_qty']) == False:
                            differences.append(f"Issue Qty: {odoo_row['issue_qty']} vs {manual_row['issue_qty']}")
                        if self.compare_numeric(odoo_row['issue_value'], manual_row['issue_value']) == False:
                            differences.append(f"Issue Value: {odoo_row['issue_value']} vs {manual_row['issue_value']}")
                        if self.compare_numeric(odoo_row['closing_qty'], manual_row['closing_qty']) == False:
                            differences.append(f"Closing Qty: {odoo_row['closing_qty']} vs {manual_row['closing_qty']}")
                        if self.compare_numeric(odoo_row['closing_value'], manual_row['closing_value']) == False:
                            differences.append(f"Closing Value: {odoo_row['closing_value']} vs {manual_row['closing_value']}")
                        
                        if differences:
                            name_matches_but_different.append({
                                'odoo': odoo_row,
                                'manual': manual_row,
                                'differences': differences
                            })
                        break
                
                if not name_match_found:
                    unmatched_odoo.append(odoo_row)
        
        # Find unmatched manual records
        for manual_row in manual_data:
            manual_code_norm = self.normalize_product_code(manual_row['product_code'])
            manual_name = manual_row['items_name']
            
            is_matched = False
            for match in matches:
                if (self.normalize_product_code(match['product_code']) == manual_code_norm and
                    match['name'] == manual_name):
                    is_matched = True
                    break
            
            if not is_matched:
                # Check if name exists in Odoo
                name_exists = any(odoo_row['product_name'] == manual_name for odoo_row in odoo_data)
                if not name_exists:
                    unmatched_manual.append(manual_row)
        
        # Write report
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("MATCH ANALYSIS REPORT\n")
            f.write("=" * 80 + "\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            f.write(f"SUMMARY\n")
            f.write("-" * 80 + "\n")
            f.write(f"Total Odoo records: {len(odoo_data)}\n")
            f.write(f"Total Manual records: {len(manual_data)}\n")
            f.write(f"Full matches (all criteria): {len(matches)}\n")
            f.write(f"  - RM matches: {len(rm_matches)}\n")
            f.write(f"  - Consumable matches: {len(consumable_matches)}\n")
            f.write(f"Name matches but other differences: {len(name_matches_but_different)}\n")
            f.write(f"Unmatched Odoo records (no name match): {len(unmatched_odoo)}\n")
            f.write(f"Unmatched Manual records (no name match): {len(unmatched_manual)}\n\n")
            
            f.write("=" * 80 + "\n")
            f.write("RECORDS WITH MATCHING NAMES BUT DIFFERENT VALUES\n")
            f.write("=" * 80 + "\n\n")
            
            # Group by difference type
            code_diff = [x for x in name_matches_but_different if any('Product Code' in d for d in x['differences'])]
            unit_diff = [x for x in name_matches_but_different if any('Unit' in d for d in x['differences'])]
            qty_value_diff = [x for x in name_matches_but_different if any(('Qty' in d or 'Value' in d) for d in x['differences'])]
            
            f.write(f"Product Code differences: {len(code_diff)}\n")
            f.write(f"Unit differences: {len(unit_diff)}\n")
            f.write(f"Quantity/Value differences: {len(qty_value_diff)}\n\n")
            
            # Detailed list
            for i, item in enumerate(name_matches_but_different, 1):
                odoo = item['odoo']
                manual = item['manual']
                f.write(f"{i}. Item Name: '{odoo['product_name']}'\n")
                f.write(f"   Odoo (Row {odoo['row_num']}): Code='{odoo['product_code']}', Unit='{odoo['unit']}', ")
                f.write(f"Opening={odoo['opening_qty']}/{odoo['opening_value']}, Closing={odoo['closing_qty']}/{odoo['closing_value']}\n")
                f.write(f"   Manual (Row {manual['row_num']}): Code='{manual['product_code']}', Unit='{manual['unit']}', ")
                f.write(f"Opening={manual['opening_qty']}/{manual['opening_value']}, Closing={manual['closing_qty']}/{manual['closing_value']}\n")
                f.write(f"   Differences: {', '.join(item['differences'])}\n\n")
            
            f.write("\n" + "=" * 80 + "\n")
            f.write("UNMATCHED ODOO RECORDS (No matching name in Manual)\n")
            f.write("=" * 80 + "\n\n")
            for i, row in enumerate(unmatched_odoo[:100], 1):  # Limit to first 100
                f.write(f"{i}. Row {row['row_num']}: Code='{row['product_code']}', Name='{row['product_name'][:60]}...', ")
                f.write(f"Unit='{row['unit']}', Opening={row['opening_qty']}/{row['opening_value']}\n")
            if len(unmatched_odoo) > 100:
                f.write(f"\n... and {len(unmatched_odoo) - 100} more unmatched Odoo records\n")
            
            f.write("\n" + "=" * 80 + "\n")
            f.write("UNMATCHED MANUAL RECORDS (No matching name in Odoo)\n")
            f.write("=" * 80 + "\n\n")
            for i, row in enumerate(unmatched_manual[:100], 1):  # Limit to first 100
                f.write(f"{i}. Row {row['row_num']}: Code='{row['product_code']}', Name='{row['items_name'][:60]}...', ")
                f.write(f"Unit='{row['unit']}', Opening={row['opening_qty']}/{row['opening_value']}\n")
            if len(unmatched_manual) > 100:
                f.write(f"\n... and {len(unmatched_manual) - 100} more unmatched Manual records\n")
        
        self.log_status(f"Analysis report saved to: {report_path}")

def main():
    root = tk.Tk()
    app = StockReportValidator(root)
    root.mainloop()

if __name__ == '__main__':
    main()

