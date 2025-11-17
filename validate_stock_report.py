"""
Stock Report Validator
Matches Product Code and Items Name between Odoo and Manual reports
Highlights matches and adds match IDs
"""

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import re

# File paths
ODOO_FILE = 'resources/Detailed Stock Report.xlsx'
MANUAL_FILE = 'resources/Monthly Stock report Aug-25.xlsx'

# Color for highlighting matches
RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

def is_data_row_odoo(row_vals):
    """Check if a row is a data row in Odoo file (has numeric SL No)"""
    if not row_vals or len(row_vals) == 0:
        return False
    sl_val = str(row_vals[0]).strip() if row_vals[0] else ''
    return bool(re.match(r'^\d+$', sl_val))

def is_data_row_manual(row_vals):
    """Check if a row is a data row in Manual file (has numeric SL # and actual data)"""
    if not row_vals or len(row_vals) < 4:
        return False
    sl_val = str(row_vals[0]).strip() if row_vals[0] else ''
    # Check if SL # is numeric
    if not re.match(r'^\d+$', sl_val):
        return False
    # Check if there's actual data (Product Code or Items Name should not be empty)
    product_code = str(row_vals[1]).strip() if len(row_vals) > 1 and row_vals[1] else ''
    items_name = str(row_vals[3]).strip() if len(row_vals) > 3 and row_vals[3] else ''
    return bool(product_code or items_name)

def normalize_text(text):
    """Normalize text for comparison (strip whitespace, handle None)"""
    if text is None:
        return ''
    return str(text).strip()

def read_odoo_data():
    """Read all data rows from Odoo file"""
    wb = openpyxl.load_workbook(ODOO_FILE)
    ws = wb.active
    
    data_rows = []
    
    # Find all header rows (page breaks) - check for "SL\nNo" in first column
    header_rows = [5]
    for i in range(6, ws.max_row + 1):
        first_cell = str(ws.cell(i, 1).value or '').strip()
        if first_cell == 'SL\nNo' or first_cell == 'SL No':
            header_rows.append(i)
    
    # Read data from each page
    for page_idx in range(len(header_rows)):
        header_start = header_rows[page_idx]
        header_next = header_rows[page_idx + 1] if page_idx + 1 < len(header_rows) else ws.max_row + 1
        
        for row_num in range(header_start + 1, header_next):
            # Read row values
            row_vals = [cell.value for cell in ws[row_num]]
            
            if is_data_row_odoo(row_vals):
                # Extract Product Code (Col 2) and Product Name (Col 3)
                product_code = normalize_text(row_vals[1] if len(row_vals) > 1 else '')
                product_name = normalize_text(row_vals[2] if len(row_vals) > 2 else '')
                
                if product_code or product_name:  # Only add if has data
                    data_rows.append({
                        'row_num': row_num,
                        'product_code': product_code,
                        'product_name': product_name
                    })
    
    wb.close()
    return data_rows

def read_manual_rm_data():
    """Read all data rows from Manual RM sheet"""
    wb = openpyxl.load_workbook(MANUAL_FILE)
    ws = wb['RM']
    
    data_rows = []
    
    # Read data rows (starting from row 6)
    for row_num in range(6, ws.max_row + 1):
        # Read columns A through N (14 columns)
        row_vals = [ws.cell(row_num, col).value for col in range(1, 15)]
        
        if is_data_row_manual(row_vals):
            # Extract Product Code (Col 2) and Items Name (Col 4)
            product_code = normalize_text(row_vals[1] if len(row_vals) > 1 else '')
            items_name = normalize_text(row_vals[3] if len(row_vals) > 3 else '')
            
            if product_code or items_name:  # Only add if has data
                data_rows.append({
                    'row_num': row_num,
                    'product_code': product_code,
                    'items_name': items_name
                })
    
    wb.close()
    return data_rows

def find_matches(odoo_data, manual_data):
    """Find matching rows based on Product Code and Name"""
    matches = []
    match_id_counter = 1
    
    for odoo_row in odoo_data:
        for manual_row in manual_data:
            # Match on Product Code AND (Product Name = Items Name)
            if (odoo_row['product_code'] and manual_row['product_code'] and 
                odoo_row['product_code'] == manual_row['product_code'] and
                odoo_row['product_name'] and manual_row['items_name'] and
                odoo_row['product_name'] == manual_row['items_name']):
                
                match_id = f'M{str(match_id_counter).zfill(4)}'
                matches.append({
                    'match_id': match_id,
                    'odoo_row_num': odoo_row['row_num'],
                    'manual_row_num': manual_row['row_num'],
                    'product_code': odoo_row['product_code'],
                    'name': odoo_row['product_name']
                })
                match_id_counter += 1
                break  # Found match, move to next odoo row
    
    return matches

def process_files():
    """Main function to process files"""
    print("Reading Odoo file...")
    odoo_data = read_odoo_data()
    print(f"Found {len(odoo_data)} data rows in Odoo file")
    
    print("Reading Manual RM sheet...")
    manual_data = read_manual_rm_data()
    print(f"Found {len(manual_data)} data rows in Manual RM sheet")
    
    print("Finding matches...")
    matches = find_matches(odoo_data, manual_data)
    print(f"Found {len(matches)} matches")
    
    if not matches:
        print("No matches found. Exiting.")
        return
    
    # Load files for modification
    print("Loading files for modification...")
    odoo_wb = openpyxl.load_workbook(ODOO_FILE)
    odoo_ws = odoo_wb.active
    
    manual_wb = openpyxl.load_workbook(MANUAL_FILE)
    manual_ws = manual_wb['RM']
    
    # Insert new first column in Odoo file
    print("Inserting match ID column in Odoo file...")
    odoo_ws.insert_cols(1)
    odoo_ws.cell(5, 1).value = 'Match ID'  # Header row
    
    # Insert new first column in Manual file
    print("Inserting match ID column in Manual file...")
    manual_ws.insert_cols(1)
    manual_ws.cell(5, 1).value = 'Match ID'  # Header row
    
    # Apply match IDs and highlight rows
    print("Applying match IDs and highlighting rows...")
    for match in matches:
        # Odoo file
        odoo_row = match['odoo_row_num']
        odoo_ws.cell(odoo_row, 1).value = match['match_id']
        # Highlight entire row (adjust column range as needed)
        for col in range(1, odoo_ws.max_column + 1):
            odoo_ws.cell(odoo_row, col).fill = RED_FILL
        
        # Manual file
        manual_row = match['manual_row_num']
        manual_ws.cell(manual_row, 1).value = match['match_id']
        # Highlight entire row (Match ID column + original columns A through N = columns 1-15)
        for col in range(1, 16):  # Columns 1-15 (Match ID + original A-N)
            manual_ws.cell(manual_row, col).fill = RED_FILL
    
    # Save files
    print("Saving Odoo file...")
    odoo_wb.save(ODOO_FILE)
    
    print("Saving Manual file...")
    manual_wb.save(MANUAL_FILE)
    
    print(f"\nCompleted! Found {len(matches)} matches.")
    print("Files have been updated with match IDs and highlighted rows.")
    print("\nMatch Summary:")
    for match in matches[:10]:  # Show first 10 matches
        print(f"  {match['match_id']}: Product Code='{match['product_code']}', Name='{match['name'][:50]}...'")
    if len(matches) > 10:
        print(f"  ... and {len(matches) - 10} more matches")

if __name__ == '__main__':
    process_files()

