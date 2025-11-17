"""
Analyze why matches are reduced - find records where Item names match but other values differ
"""

import openpyxl
import re

ODOO_FILE = 'resources/Detailed Stock Report.xlsx'
MANUAL_FILE = 'resources/Monthly Stock report Aug-25.xlsx'

def normalize_text(text):
    if text is None:
        return ''
    return str(text).strip()

def normalize_numeric(text):
    """Normalize numeric values for comparison"""
    if text is None:
        return 0.0
    if isinstance(text, (int, float)):
        return round(float(text), 2)
    text_str = str(text).strip()
    if text_str == '' or text_str.lower() == 'none' or text_str == '-':
        return 0.0
    if text_str.startswith('(') and text_str.endswith(')'):
        text_str = text_str[1:-1].strip()
        is_negative = True
    else:
        is_negative = False
    text_str = text_str.replace(' ', '').replace(',', '').replace('$', '').replace('₹', '').replace('€', '').replace('£', '')
    if text_str == '':
        return 0.0
    try:
        num_value = float(text_str)
        if is_negative:
            num_value = -num_value
        return round(num_value, 2)
    except (ValueError, TypeError):
        return 0.0

def read_odoo_data():
    """Read Odoo data"""
    wb = openpyxl.load_workbook(ODOO_FILE)
    ws = wb.active
    has_match_id = str(ws.cell(5, 1).value or '').strip() == 'Match ID'
    
    data_rows = []
    header_rows = [5]
    sl_col = 2 if has_match_id else 1
    
    for i in range(6, ws.max_row + 1):
        first_cell = str(ws.cell(i, sl_col).value or '').strip()
        if first_cell == 'SL\nNo' or first_cell == 'SL No' or first_cell == 'SL':
            header_rows.append(i)
    
    if has_match_id:
        code_col, name_col, unit_col = 3, 4, 6
        opening_qty_col, opening_value_col = 7, 8
        receive_qty_col, receive_value_col = 9, 10
        issue_qty_col, issue_value_col = 11, 12
        closing_qty_col, closing_value_col = 13, 14
    else:
        code_col, name_col, unit_col = 2, 3, 5
        opening_qty_col, opening_value_col = 6, 7
        receive_qty_col, receive_value_col = 8, 9
        issue_qty_col, issue_value_col = 10, 11
        closing_qty_col, closing_value_col = 12, 13
    
    for page_idx in range(len(header_rows)):
        header_start = header_rows[page_idx]
        header_next = header_rows[page_idx + 1] if page_idx + 1 < len(header_rows) else ws.max_row + 1
        
        for row_num in range(header_start + 1, header_next):
            sl_val = ws.cell(row_num, sl_col).value
            if sl_val is None:
                continue
            sl_str = str(sl_val).strip()
            if not (sl_str.isdigit() or (isinstance(sl_val, (int, float)))):
                continue
            
            product_code = normalize_text(ws.cell(row_num, code_col).value)
            product_name = normalize_text(ws.cell(row_num, name_col).value)
            unit = normalize_text(ws.cell(row_num, unit_col).value)
            opening_qty = normalize_numeric(ws.cell(row_num, opening_qty_col).value)
            opening_value = normalize_numeric(ws.cell(row_num, opening_value_col).value)
            receive_qty = normalize_numeric(ws.cell(row_num, receive_qty_col).value)
            receive_value = normalize_numeric(ws.cell(row_num, receive_value_col).value)
            issue_qty = normalize_numeric(ws.cell(row_num, issue_qty_col).value)
            issue_value = normalize_numeric(ws.cell(row_num, issue_value_col).value)
            closing_qty = normalize_numeric(ws.cell(row_num, closing_qty_col).value)
            closing_value = normalize_numeric(ws.cell(row_num, closing_value_col).value)
            
            if product_code or product_name:
                data_rows.append({
                    'row_num': row_num,
                    'product_code': product_code,
                    'product_name': product_name,
                    'unit': unit,
                    'opening_qty': opening_qty,
                    'opening_value': opening_value,
                    'receive_qty': receive_qty,
                    'receive_value': receive_value,
                    'issue_qty': issue_qty,
                    'issue_value': issue_value,
                    'closing_qty': closing_qty,
                    'closing_value': closing_value
                })
    
    wb.close()
    return data_rows

def read_manual_data():
    """Read Manual RM data"""
    wb_check = openpyxl.load_workbook(MANUAL_FILE, data_only=True)
    ws_check = wb_check['RM']
    has_match_id = str(ws_check.cell(5, 1).value or '').strip() == 'Match ID'
    wb_check.close()
    
    wb = openpyxl.load_workbook(MANUAL_FILE, data_only=True)
    ws = wb['RM']
    
    data_rows = []
    
    if has_match_id:
        sl_col, code_col, name_col, unit_col = 2, 3, 5, 6
        opening_qty_col, opening_value_col = 7, 8
        receive_qty_col, receive_value_col = 9, 10
        issue_qty_col, issue_value_col = 11, 12
        closing_qty_col, closing_value_col = 13, 14
    else:
        sl_col, code_col, name_col, unit_col = 1, 2, 4, 5
        opening_qty_col, opening_value_col = 6, 7
        receive_qty_col, receive_value_col = 8, 9
        issue_qty_col, issue_value_col = 10, 11
        closing_qty_col, closing_value_col = 12, 13
    
    for row_num in range(6, ws.max_row + 1):
        sl_val = ws.cell(row_num, sl_col).value
        if sl_val is None:
            continue
        sl_str = str(sl_val).strip()
        if not (sl_str.isdigit() or (isinstance(sl_val, (int, float)))):
            continue
        
        product_code = normalize_text(ws.cell(row_num, code_col).value)
        items_name = normalize_text(ws.cell(row_num, name_col).value)
        unit = normalize_text(ws.cell(row_num, unit_col).value)
        opening_qty = normalize_numeric(ws.cell(row_num, opening_qty_col).value)
        opening_value = normalize_numeric(ws.cell(row_num, opening_value_col).value)
        receive_qty = normalize_numeric(ws.cell(row_num, receive_qty_col).value)
        receive_value = normalize_numeric(ws.cell(row_num, receive_value_col).value)
        issue_qty = normalize_numeric(ws.cell(row_num, issue_qty_col).value)
        issue_value = normalize_numeric(ws.cell(row_num, issue_value_col).value)
        closing_qty = normalize_numeric(ws.cell(row_num, closing_qty_col).value)
        closing_value = normalize_numeric(ws.cell(row_num, closing_value_col).value)
        
        if product_code or items_name:
            data_rows.append({
                'row_num': row_num,
                'product_code': product_code,
                'items_name': items_name,
                'unit': unit,
                'opening_qty': opening_qty,
                'opening_value': opening_value,
                'receive_qty': receive_qty,
                'receive_value': receive_value,
                'issue_qty': issue_qty,
                'issue_value': issue_value,
                'closing_qty': closing_qty,
                'closing_value': closing_value
            })
    
    wb.close()
    return data_rows

def analyze():
    print("=" * 80)
    print("ANALYSIS: Why matches are reduced")
    print("=" * 80)
    
    print("\nReading files...")
    odoo_data = read_odoo_data()
    manual_data = read_manual_data()
    
    print(f"Odoo file: {len(odoo_data)} records")
    print(f"Manual file: {len(manual_data)} records")
    
    # Find matches (all criteria)
    print("\nFinding full matches (all criteria)...")
    full_matches = []
    for odoo_row in odoo_data:
        for manual_row in manual_data:
            if (odoo_row['product_code'] and manual_row['product_code'] and 
                odoo_row['product_code'] == manual_row['product_code'] and
                odoo_row['product_name'] and manual_row['items_name'] and
                odoo_row['product_name'] == manual_row['items_name'] and
                odoo_row['unit'] and manual_row['unit'] and
                odoo_row['unit'] == manual_row['unit'] and
                odoo_row['opening_qty'] == manual_row['opening_qty'] and
                odoo_row['opening_value'] == manual_row['opening_value'] and
                odoo_row['receive_qty'] == manual_row['receive_qty'] and
                odoo_row['receive_value'] == manual_row['receive_value'] and
                odoo_row['issue_qty'] == manual_row['issue_qty'] and
                odoo_row['issue_value'] == manual_row['issue_value'] and
                odoo_row['closing_qty'] == manual_row['closing_qty'] and
                odoo_row['closing_value'] == manual_row['closing_value']):
                full_matches.append((odoo_row, manual_row))
                break
    
    print(f"Full matches found: {len(full_matches)}")
    
    # Find name matches but other differences
    print("\nFinding records where Item Names match but other values differ...")
    name_matches_but_different = []
    
    for odoo_row in odoo_data:
        for manual_row in manual_data:
            # Check if names match
            if (odoo_row['product_name'] and manual_row['items_name'] and
                odoo_row['product_name'] == manual_row['items_name']):
                
                # Check what's different
                differences = []
                if odoo_row['product_code'] != manual_row['product_code']:
                    differences.append(f"Product Code: '{odoo_row['product_code']}' vs '{manual_row['product_code']}'")
                if odoo_row['unit'] != manual_row['unit']:
                    differences.append(f"Unit: '{odoo_row['unit']}' vs '{manual_row['unit']}'")
                if odoo_row['opening_qty'] != manual_row['opening_qty']:
                    differences.append(f"Opening Qty: {odoo_row['opening_qty']} vs {manual_row['opening_qty']}")
                if odoo_row['opening_value'] != manual_row['opening_value']:
                    differences.append(f"Opening Value: {odoo_row['opening_value']} vs {manual_row['opening_value']}")
                if odoo_row['receive_qty'] != manual_row['receive_qty']:
                    differences.append(f"Receive Qty: {odoo_row['receive_qty']} vs {manual_row['receive_qty']}")
                if odoo_row['receive_value'] != manual_row['receive_value']:
                    differences.append(f"Receive Value: {odoo_row['receive_value']} vs {manual_row['receive_value']}")
                if odoo_row['issue_qty'] != manual_row['issue_qty']:
                    differences.append(f"Issue Qty: {odoo_row['issue_qty']} vs {manual_row['issue_qty']}")
                if odoo_row['issue_value'] != manual_row['issue_value']:
                    differences.append(f"Issue Value: {odoo_row['issue_value']} vs {manual_row['issue_value']}")
                if odoo_row['closing_qty'] != manual_row['closing_qty']:
                    differences.append(f"Closing Qty: {odoo_row['closing_qty']} vs {manual_row['closing_qty']}")
                if odoo_row['closing_value'] != manual_row['closing_value']:
                    differences.append(f"Closing Value: {odoo_row['closing_value']} vs {manual_row['closing_value']}")
                
                if differences:
                    name_matches_but_different.append({
                        'odoo': odoo_row,
                        'manual': manual_row,
                        'differences': differences
                    })
    
    print(f"\nRecords with matching Item Names but different values: {len(name_matches_but_different)}")
    
    # Group by difference type
    print("\n" + "=" * 80)
    print("ANALYSIS BY DIFFERENCE TYPE")
    print("=" * 80)
    
    code_diff = []
    unit_diff = []
    qty_value_diff = []
    
    for item in name_matches_but_different:
        odoo = item['odoo']
        manual = item['manual']
        
        if odoo['product_code'] != manual['product_code']:
            code_diff.append(item)
        if odoo['unit'] != manual['unit']:
            unit_diff.append(item)
        if (odoo['opening_qty'] != manual['opening_qty'] or
            odoo['opening_value'] != manual['opening_value'] or
            odoo['receive_qty'] != manual['receive_qty'] or
            odoo['receive_value'] != manual['receive_value'] or
            odoo['issue_qty'] != manual['issue_qty'] or
            odoo['issue_value'] != manual['issue_value'] or
            odoo['closing_qty'] != manual['closing_qty'] or
            odoo['closing_value'] != manual['closing_value']):
            qty_value_diff.append(item)
    
    print(f"\n1. Product Code differences: {len(code_diff)}")
    print(f"2. Unit differences: {len(unit_diff)}")
    print(f"3. Quantity/Value differences: {len(qty_value_diff)}")
    
    # Show examples
    print("\n" + "=" * 80)
    print("EXAMPLES: Product Name matches but Product Code differs")
    print("=" * 80)
    for i, item in enumerate(code_diff[:10], 1):
        odoo = item['odoo']
        manual = item['manual']
        print(f"\n{i}. Item Name: '{odoo['product_name']}'")
        print(f"   Odoo (Row {odoo['row_num']}): Code='{odoo['product_code']}', Unit='{odoo['unit']}', Opening Qty={odoo['opening_qty']}, Opening Value={odoo['opening_value']}")
        print(f"   Manual (Row {manual['row_num']}): Code='{manual['product_code']}', Unit='{manual['unit']}', Opening Qty={manual['opening_qty']}, Opening Value={manual['opening_value']}")
        print(f"   Differences: {', '.join(item['differences'][:3])}")
    
    print("\n" + "=" * 80)
    print("EXAMPLES: Product Name matches but Quantity/Value differs")
    print("=" * 80)
    for i, item in enumerate(qty_value_diff[:10], 1):
        odoo = item['odoo']
        manual = item['manual']
        print(f"\n{i}. Item Name: '{odoo['product_name']}'")
        print(f"   Odoo (Row {odoo['row_num']}): Code='{odoo['product_code']}', Opening Qty={odoo['opening_qty']}, Opening Value={odoo['opening_value']}, Closing Qty={odoo['closing_qty']}, Closing Value={odoo['closing_value']}")
        print(f"   Manual (Row {manual['row_num']}): Code='{manual['product_code']}', Opening Qty={manual['opening_qty']}, Opening Value={manual['opening_value']}, Closing Qty={manual['closing_qty']}, Closing Value={manual['closing_value']}")
        print(f"   Differences: {', '.join(item['differences'][:5])}")
    
    # Summary statistics
    print("\n" + "=" * 80)
    print("SUMMARY STATISTICS")
    print("=" * 80)
    print(f"Total Odoo records: {len(odoo_data)}")
    print(f"Total Manual records: {len(manual_data)}")
    print(f"Full matches (all criteria): {len(full_matches)}")
    print(f"Name matches but other differences: {len(name_matches_but_different)}")
    print(f"Potential matches if criteria relaxed: {len(full_matches) + len(name_matches_but_different)}")
    
    # Save detailed report
    print("\n" + "=" * 80)
    print("Saving detailed analysis to 'match_analysis_report.txt'...")
    with open('match_analysis_report.txt', 'w', encoding='utf-8') as f:
        f.write("MATCH ANALYSIS REPORT\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Total Odoo records: {len(odoo_data)}\n")
        f.write(f"Total Manual records: {len(manual_data)}\n")
        f.write(f"Full matches (all criteria): {len(full_matches)}\n")
        f.write(f"Name matches but other differences: {len(name_matches_but_different)}\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("DETAILED: Name matches but Product Code differs\n")
        f.write("=" * 80 + "\n")
        for item in code_diff:
            odoo = item['odoo']
            manual = item['manual']
            f.write(f"\nItem Name: '{odoo['product_name']}'\n")
            f.write(f"  Odoo Row {odoo['row_num']}: Code='{odoo['product_code']}', Unit='{odoo['unit']}'\n")
            f.write(f"  Manual Row {manual['row_num']}: Code='{manual['product_code']}', Unit='{manual['unit']}'\n")
            f.write(f"  All Differences: {', '.join(item['differences'])}\n")
        
        f.write("\n" + "=" * 80 + "\n")
        f.write("DETAILED: Name matches but Quantity/Value differs\n")
        f.write("=" * 80 + "\n")
        for item in qty_value_diff[:50]:  # Limit to first 50
            odoo = item['odoo']
            manual = item['manual']
            f.write(f"\nItem Name: '{odoo['product_name']}'\n")
            f.write(f"  Odoo Row {odoo['row_num']}: Code='{odoo['product_code']}', Opening={odoo['opening_qty']}/{odoo['opening_value']}, Closing={odoo['closing_qty']}/{odoo['closing_value']}\n")
            f.write(f"  Manual Row {manual['row_num']}: Code='{manual['product_code']}', Opening={manual['opening_qty']}/{manual['opening_value']}, Closing={manual['closing_qty']}/{manual['closing_value']}\n")
            f.write(f"  Differences: {', '.join(item['differences'])}\n")
    
    print("Analysis complete! Report saved to 'match_analysis_report.txt'")

if __name__ == '__main__':
    analyze()

